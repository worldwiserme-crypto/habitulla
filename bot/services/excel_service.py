"""Excel report generator — peach/coral design matching habit tracker template.

Two sheets with unified aesthetic:
  • "Odatlar" — 31-day habit grid + daily/weekly charts
  • "Budjet"  — categorized income/expenses with formulas

Design inspired by popular habit-tracker Google Sheets templates.
All text in Uzbek.
"""
from __future__ import annotations

import asyncio
import os
import tempfile
from calendar import monthrange
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from bot.services.db_service import db
from bot.utils.formatters import UZ_MONTHS, UZ_WEEKDAYS
from bot.utils.logger import logger

# ════════════════════════════════════════════════════════════════
# PALETTE — peach/coral theme (matches habit tracker template)
# ════════════════════════════════════════════════════════════════
PEACH_DARK = "D97D54"       # Coral — headers, filled checkboxes
PEACH_MID = "E8A87C"        # Primary peach — banners
PEACH_LIGHT = "F5C9A6"      # Light peach — progress bars
PEACH_BG = "FCE4D6"         # Very light peach — alt rows
PEACH_CREAM = "FDF2E9"      # Cream — page background
INK = "3D2817"              # Dark brown — text
MUTED = "8B7355"            # Muted brown — secondary text
GREEN = "7EAE76"            # Balance positive
RED = "C65D54"              # Balance negative
WHITE = "FFFFFF"

# ── Fonts ──
FONT_TITLE = Font(name="Calibri", size=24, bold=True, color=INK)
FONT_BANNER = Font(name="Calibri", size=14, bold=True, color=WHITE)
FONT_H1 = Font(name="Calibri", size=13, bold=True, color=INK)
FONT_H2 = Font(name="Calibri", size=11, bold=True, color=WHITE)
FONT_BODY = Font(name="Calibri", size=10, color=INK)
FONT_BOLD = Font(name="Calibri", size=10, bold=True, color=INK)
FONT_SMALL = Font(name="Calibri", size=9, color=MUTED, italic=True)
FONT_NUM = Font(name="Calibri", size=10, color=INK)
FONT_CHECK = Font(name="Calibri", size=11, bold=True, color=WHITE)

# ── Fills ──
FILL_BANNER = PatternFill("solid", fgColor=PEACH_MID)
FILL_HEADER = PatternFill("solid", fgColor=PEACH_DARK)
FILL_LIGHT = PatternFill("solid", fgColor=PEACH_BG)
FILL_CREAM = PatternFill("solid", fgColor=PEACH_CREAM)
FILL_CHECK = PatternFill("solid", fgColor=PEACH_DARK)
FILL_EMPTY = PatternFill("solid", fgColor=PEACH_CREAM)
FILL_PROGRESS = PatternFill("solid", fgColor=PEACH_LIGHT)
FILL_GREEN = PatternFill("solid", fgColor=GREEN)
FILL_RED = PatternFill("solid", fgColor=RED)
FILL_WHITE = PatternFill("solid", fgColor=WHITE)

# ── Borders ──
_BORDER_THIN = Side(border_style="thin", color=PEACH_LIGHT)
_BORDER_WHITE = Side(border_style="medium", color=WHITE)
BORDER_ALL = Border(left=_BORDER_THIN, right=_BORDER_THIN, top=_BORDER_THIN, bottom=_BORDER_THIN)
BORDER_WHITE = Border(left=_BORDER_WHITE, right=_BORDER_WHITE, top=_BORDER_WHITE, bottom=_BORDER_WHITE)

# ── Alignments ──
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", indent=1)
RIGHT = Alignment(horizontal="right", vertical="center", indent=1)


# ════════════════════════════════════════════════════════════════
# BUDGET CATEGORY STRUCTURE (Vertex42-inspired, in Uzbek)
# ════════════════════════════════════════════════════════════════
BUDGET_CATEGORIES = {
    "KIRIM": [
        ("Oylik maosh", "salary"),
        ("Qo'shimcha ish", "freelance"),
        ("Foiz daromadi", "interest"),
        ("Dividend", "dividend"),
        ("Olingan sovg'alar", "gift"),
        ("Qaytarilgan pul", "refund"),
        ("Jamg'armadan", "from_savings"),
        ("Boshqa", "other_income"),
    ],
    "UY XARAJATLARI": [
        ("Ijara / Ipoteka", "rent"),
        ("Uy sug'urtasi", "home_insurance"),
        ("Elektr", "electricity"),
        ("Gaz", "gas"),
        ("Suv / Kanalizatsiya", "water"),
        ("Telefon", "phone"),
        ("Internet", "internet"),
        ("Kommunal - boshqa", "utility_other"),
        ("Mebel / Jihozlar", "furniture"),
        ("Ta'mirlash", "repair"),
        ("Boshqa", "home_other"),
    ],
    "TRANSPORT": [
        ("Avto to'lov", "car_payment"),
        ("Avto sug'urta", "car_insurance"),
        ("Yoqilg'i", "fuel"),
        ("Taksi / Avtobus / Metro", "transit"),
        ("Avto ta'mirlash", "car_repair"),
        ("Avto ro'yxati", "car_registration"),
        ("Boshqa", "transport_other"),
    ],
    "KUNDALIK": [
        ("Oziq-ovqat", "groceries"),
        ("Restoran / Kafe", "dining"),
        ("Kiyim", "clothing"),
        ("Shaxsiy buyumlar", "personal_supplies"),
        ("Tozalash vositalari", "cleaning"),
        ("Salon / Sartaroshxona", "salon"),
        ("Ta'lim / Kurslar", "education"),
        ("Soglik / Dori", "health"),
        ("Uy hayvonlari", "pets"),
        ("Boshqa", "daily_other"),
    ],
    "KO'NGIL OCHAR": [
        ("Kino / Teatr / Konsert", "entertainment"),
        ("Sayohat / Ta'til", "travel"),
        ("Kitoblar", "books"),
        ("Sport / Zal", "sports"),
        ("Obunalar (Netflix va h.k.)", "subscriptions"),
        ("Hobi", "hobby"),
        ("O'yinlar", "games"),
        ("Boshqa", "entertainment_other"),
    ],
    "JAMG'ARMA": [
        ("Favqulodda jamg'arma", "emergency"),
        ("Pensiya jamg'armasi", "retirement"),
        ("Investitsiya", "investment"),
        ("Avto uchun jamg'arma", "car_savings"),
        ("Boshqa maqsad", "other_savings"),
    ],
}

# Map user-facing category names (from ai_service) to our budget categories
CATEGORY_MAP = {
    "oziq-ovqat": ("KUNDALIK", "Oziq-ovqat"),
    "transport": ("TRANSPORT", "Taksi / Avtobus / Metro"),
    "soglik": ("KUNDALIK", "Soglik / Dori"),
    "kiyim": ("KUNDALIK", "Kiyim"),
    "kommunal": ("UY XARAJATLARI", "Kommunal - boshqa"),
    "ta'lim": ("KUNDALIK", "Ta'lim / Kurslar"),
    "ko'ngil-ochar": ("KO'NGIL OCHAR", "Kino / Teatr / Konsert"),
    "boshqa": ("KUNDALIK", "Boshqa"),
}


# ════════════════════════════════════════════════════════════════
# HELPERS
# ════════════════════════════════════════════════════════════════
def _set_sheet_defaults(ws: Worksheet) -> None:
    """Apply cream background and default styling."""
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 110
    # Fill entire visible area with cream
    for row in range(1, 80):
        for col in range(1, 45):
            ws.cell(row=row, column=col).fill = FILL_CREAM


def _merge_and_style(
    ws: Worksheet,
    range_str: str,
    value: str,
    fill: PatternFill,
    font: Font,
    alignment: Alignment = CENTER,
) -> None:
    ws.merge_cells(range_str)
    first = range_str.split(":")[0]
    cell = ws[first]
    cell.value = value
    cell.fill = fill
    cell.font = font
    cell.alignment = alignment


def _draw_progress_bar(
    ws: Worksheet, row: int, col: int, percent: float, width: int = 1
) -> None:
    """Visual progress indicator in a cell using character-based bar."""
    cell = ws.cell(row=row, column=col)
    filled_blocks = min(10, max(0, int(percent / 10)))
    bar = "█" * filled_blocks + "░" * (10 - filled_blocks)
    cell.value = f"{bar}  {percent:.0f}%"
    cell.font = Font(name="Consolas", size=9, color=PEACH_DARK, bold=True)
    cell.alignment = LEFT
    cell.fill = FILL_CREAM


# ════════════════════════════════════════════════════════════════
# HABIT SHEET
# ════════════════════════════════════════════════════════════════
def _build_habit_sheet(
    ws: Worksheet,
    habits: List[Dict[str, Any]],
    start: date,
    end: date,
) -> None:
    """Build the 'Odatlar' sheet with 31-day grid + stats."""
    _set_sheet_defaults(ws)

    # ── 1. TITLE BANNER ───────────────────────────────────────
    ws.row_dimensions[1].height = 50
    _merge_and_style(
        ws, "A1:AI1",
        "  🎯  ODATLAR TRACKER",
        FILL_BANNER,
        Font(name="Calibri", size=22, bold=True, color=WHITE),
        Alignment(horizontal="left", vertical="center", indent=2),
    )
    ws.row_dimensions[2].height = 6  # Spacer

    # ── 2. SUBTITLE ──────────────────────────────────────────
    ws.row_dimensions[3].height = 20
    _merge_and_style(
        ws, "A3:AI3",
        f"Davr: {start.strftime('%d.%m.%Y')} — {end.strftime('%d.%m.%Y')}  ·  "
        f"Har kuni odatlaringizni kuzatib boring va maqsadlaringizga erishing",
        FILL_CREAM,
        Font(name="Calibri", size=10, italic=True, color=MUTED),
        Alignment(horizontal="left", vertical="center", indent=2),
    )
    ws.row_dimensions[4].height = 10

    # ── 3. Aggregate habit data ──────────────────────────────
    # Unique habits in period
    habit_counts: Counter = Counter()
    habit_by_date: Dict[str, set] = defaultdict(set)
    for log in habits:
        name = (log.get("habit_name") or "Noma'lum").strip()
        habit_counts[name] += 1
        logged_d = str(log.get("logged_date") or "")
        habit_by_date[logged_d].add(name)

    unique_habits = [h for h, _ in habit_counts.most_common()]
    total_days = (end - start).days + 1

    # ── 4. KUNLIK ODATLAR HEADER ─────────────────────────────
    header_row = 5
    ws.row_dimensions[header_row].height = 30

    # "KUNLIK ODATLAR" label (cols A-B)
    _merge_and_style(
        ws, f"A{header_row}:B{header_row}",
        "KUNLIK ODATLAR",
        FILL_HEADER, FONT_H2, CENTER,
    )
    _merge_and_style(
        ws, f"C{header_row}:C{header_row}",
        "JARAYON",
        FILL_HEADER, FONT_H2, CENTER,
    )
    _merge_and_style(
        ws, f"D{header_row}:D{header_row}",
        "MAQSAD",
        FILL_HEADER, FONT_H2, CENTER,
    )

    # Day columns E-AI (31 days)
    num_days = min(total_days, 31)
    for day_idx in range(num_days):
        col = 5 + day_idx  # Column E = 5
        day_num = day_idx + 1
        cell = ws.cell(row=header_row, column=col)
        cell.value = day_num
        cell.fill = FILL_HEADER
        cell.font = FONT_H2
        cell.alignment = CENTER

    done_col = 5 + num_days
    cell = ws.cell(row=header_row, column=done_col)
    cell.value = "✓"
    cell.fill = FILL_HEADER
    cell.font = FONT_H2
    cell.alignment = CENTER

    # ── 5. HABIT ROWS ────────────────────────────────────────
    current_row = header_row + 1

    if not unique_habits:
        ws.row_dimensions[current_row].height = 40
        _merge_and_style(
            ws, f"A{current_row}:{get_column_letter(done_col)}{current_row}",
            "📭 Ushbu davrda odatlar ro'yxatga olinmagan. /start ni bosib boshlang!",
            FILL_LIGHT,
            Font(name="Calibri", size=11, italic=True, color=MUTED),
            CENTER,
        )
        current_row += 2
    else:
        goal = total_days  # Default goal = all days
        for idx, habit_name in enumerate(unique_habits):
            row = current_row
            ws.row_dimensions[row].height = 24

            # Alternate row background
            row_fill = FILL_LIGHT if idx % 2 == 0 else FILL_CREAM

            # A-B: Habit name (merged)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            name_cell = ws.cell(row=row, column=1)
            name_cell.value = f"  {habit_name}"
            name_cell.font = FONT_BODY
            name_cell.alignment = LEFT
            name_cell.fill = row_fill
            # Fill merge target too
            ws.cell(row=row, column=2).fill = row_fill

            # C: Progress bar
            done_count = habit_counts[habit_name]
            percent = (done_count / goal * 100) if goal else 0
            prog_cell = ws.cell(row=row, column=3)
            bar_blocks = min(10, max(0, int(percent / 10)))
            prog_cell.value = "█" * bar_blocks + "░" * (10 - bar_blocks)
            prog_cell.font = Font(name="Consolas", size=10, color=PEACH_DARK, bold=True)
            prog_cell.alignment = CENTER
            prog_cell.fill = row_fill

            # D: Goal
            goal_cell = ws.cell(row=row, column=4)
            goal_cell.value = goal
            goal_cell.font = FONT_BOLD
            goal_cell.alignment = CENTER
            goal_cell.fill = row_fill

            # E-AI: Day checkboxes
            for day_idx in range(num_days):
                col = 5 + day_idx
                check_date = start + timedelta(days=day_idx)
                checked = habit_name in habit_by_date.get(check_date.isoformat(), set())

                c = ws.cell(row=row, column=col)
                if checked:
                    c.value = "✓"
                    c.fill = FILL_CHECK
                    c.font = FONT_CHECK
                else:
                    c.value = ""
                    c.fill = FILL_EMPTY
                    c.font = FONT_BODY
                c.alignment = CENTER
                c.border = BORDER_WHITE

            # Done count
            done_cell = ws.cell(row=row, column=done_col)
            done_cell.value = done_count
            done_cell.font = FONT_BOLD
            done_cell.alignment = CENTER
            done_cell.fill = row_fill

            current_row += 1

    current_row += 2  # Spacer

    # ── 6. STATISTICS PANEL ──────────────────────────────────
    ws.row_dimensions[current_row].height = 30
    _merge_and_style(
        ws, f"A{current_row}:{get_column_letter(done_col)}{current_row}",
        "📊 UMUMIY STATISTIKA",
        FILL_HEADER, FONT_H2,
        Alignment(horizontal="left", vertical="center", indent=2),
    )
    current_row += 1

    total_logs = sum(habit_counts.values())
    unique_count = len(unique_habits)
    active_days = len(habit_by_date)
    consistency = (active_days / total_days * 100) if total_days else 0

    stats = [
        ("Jami yozuvlar", f"{total_logs}"),
        ("Noyob odatlar", f"{unique_count}"),
        ("Faol kunlar", f"{active_days} / {total_days}"),
        ("Izchillik", f"{consistency:.0f}%"),
    ]
    for label, val in stats:
        r = current_row
        ws.row_dimensions[r].height = 22
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        lbl = ws.cell(row=r, column=1)
        lbl.value = f"  {label}"
        lbl.font = FONT_BODY
        lbl.fill = FILL_LIGHT
        lbl.alignment = LEFT
        for fc in range(2, 5):
            ws.cell(row=r, column=fc).fill = FILL_LIGHT

        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=done_col)
        vcell = ws.cell(row=r, column=5)
        vcell.value = val
        vcell.font = FONT_BOLD
        vcell.fill = FILL_LIGHT
        vcell.alignment = LEFT
        for fc in range(6, done_col + 1):
            ws.cell(row=r, column=fc).fill = FILL_LIGHT
        current_row += 1

    # ── 7. TOP HABITS CHART ──────────────────────────────────
    if unique_habits:
        current_row += 2
        chart_start_row = current_row
        ws.row_dimensions[current_row].height = 25
        _merge_and_style(
            ws, f"A{current_row}:{get_column_letter(done_col)}{current_row}",
            "📈 TOP ODATLAR",
            FILL_HEADER, FONT_H2,
            Alignment(horizontal="left", vertical="center", indent=2),
        )
        current_row += 1

        # Data table for chart
        data_start = current_row
        header_r = current_row
        ws.cell(row=header_r, column=1, value="Odat").font = FONT_BOLD
        ws.cell(row=header_r, column=2, value="Bajarildi").font = FONT_BOLD
        ws.cell(row=header_r, column=1).fill = FILL_LIGHT
        ws.cell(row=header_r, column=2).fill = FILL_LIGHT
        current_row += 1

        for name in unique_habits[:10]:
            ws.cell(row=current_row, column=1, value=name).font = FONT_BODY
            ws.cell(row=current_row, column=2, value=habit_counts[name]).font = FONT_BOLD
            ws.cell(row=current_row, column=1).fill = FILL_CREAM
            ws.cell(row=current_row, column=2).fill = FILL_CREAM
            current_row += 1

        # Create bar chart
        chart = BarChart()
        chart.type = "bar"
        chart.style = 11
        chart.title = "Top odatlar (bajarilish soni)"
        chart.y_axis.title = None
        chart.x_axis.title = None

        data_ref = Reference(ws, min_col=2, min_row=data_start, max_row=current_row - 1, max_col=2)
        cats_ref = Reference(ws, min_col=1, min_row=data_start + 1, max_row=current_row - 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.height = 10
        chart.width = 20
        chart.legend = None

        # Place chart next to data
        chart_anchor = f"D{data_start}"
        ws.add_chart(chart, chart_anchor)

    # ── 8. COLUMN WIDTHS ──────────────────────────────────────
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 9
    for day_idx in range(num_days):
        ws.column_dimensions[get_column_letter(5 + day_idx)].width = 3.2
    ws.column_dimensions[get_column_letter(done_col)].width = 6

    # Freeze header
    ws.freeze_panes = "E6"


# ════════════════════════════════════════════════════════════════
# BUDGET SHEET
# ════════════════════════════════════════════════════════════════
def _build_budget_sheet(
    ws: Worksheet,
    budget: List[Dict[str, Any]],
    start: date,
    end: date,
    currency: str = "UZS",
) -> None:
    """Build the 'Budjet' sheet with category breakdown."""
    _set_sheet_defaults(ws)

    # ── 1. TITLE BANNER ───────────────────────────────────────
    ws.row_dimensions[1].height = 50
    _merge_and_style(
        ws, "A1:I1",
        "  💰  BUDJET TRACKER",
        FILL_BANNER,
        Font(name="Calibri", size=22, bold=True, color=WHITE),
        Alignment(horizontal="left", vertical="center", indent=2),
    )
    ws.row_dimensions[2].height = 6

    ws.row_dimensions[3].height = 20
    _merge_and_style(
        ws, "A3:I3",
        f"Davr: {start.strftime('%d.%m.%Y')} — {end.strftime('%d.%m.%Y')}  ·  "
        f"Valyuta: {currency}  ·  Daromad va xarajatlaringizni to'liq tahlil qiling",
        FILL_CREAM,
        Font(name="Calibri", size=10, italic=True, color=MUTED),
        Alignment(horizontal="left", vertical="center", indent=2),
    )
    ws.row_dimensions[4].height = 10

    # ── 2. SUMMARY CARDS (3 big cards) ───────────────────────
    total_income = sum(float(b["amount"]) for b in budget if b.get("type") == "income")
    total_expense = sum(float(b["amount"]) for b in budget if b.get("type") == "expense")
    balance = total_income - total_expense

    card_row = 5
    ws.row_dimensions[card_row].height = 25
    ws.row_dimensions[card_row + 1].height = 35
    ws.row_dimensions[card_row + 2].height = 10

    cards = [
        ("KIRIM", total_income, FILL_GREEN, "A", "C"),
        ("CHIQIM", total_expense, FILL_RED, "D", "F"),
        ("QOLDIQ", balance, FILL_GREEN if balance >= 0 else FILL_RED, "G", "I"),
    ]
    for label, value, fill, start_col, end_col in cards:
        _merge_and_style(
            ws, f"{start_col}{card_row}:{end_col}{card_row}",
            label,
            fill,
            Font(name="Calibri", size=11, bold=True, color=WHITE),
            CENTER,
        )
        _merge_and_style(
            ws, f"{start_col}{card_row + 1}:{end_col}{card_row + 1}",
            f"{value:,.0f} {currency}".replace(",", " "),
            fill,
            Font(name="Calibri", size=16, bold=True, color=WHITE),
            CENTER,
        )

    # ── 3. CATEGORY TABLE HEADERS ────────────────────────────
    current_row = card_row + 4
    ws.row_dimensions[current_row].height = 30
    _merge_and_style(ws, f"A{current_row}:C{current_row}", "KATEGORIYA", FILL_HEADER, FONT_H2)
    _merge_and_style(ws, f"D{current_row}:E{current_row}", "REJA", FILL_HEADER, FONT_H2)
    _merge_and_style(ws, f"F{current_row}:G{current_row}", "HAQIQATDA", FILL_HEADER, FONT_H2)
    _merge_and_style(ws, f"H{current_row}:I{current_row}", "FARQ", FILL_HEADER, FONT_H2)
    current_row += 1

    # ── 4. Group user's actual budget by our category structure ─
    # Map budget logs to category buckets
    actuals: Dict[Tuple[str, str], float] = defaultdict(float)
    for log in budget:
        if log.get("type") != "expense":
            # Income goes into KIRIM bucket
            actuals[("KIRIM", "Oylik maosh")] += float(log.get("amount") or 0)
            continue
        raw_cat = (log.get("category") or "boshqa").lower()
        group, sub = CATEGORY_MAP.get(raw_cat, ("KUNDALIK", "Boshqa"))
        actuals[(group, sub)] += float(log.get("amount") or 0)

    # Also put income in its proper category (not just salary)
    income_total = sum(float(b["amount"]) for b in budget if b.get("type") == "income")
    # Reset KIRIM from above loop (we assigned salary incorrectly there)
    for k in list(actuals.keys()):
        if k[0] == "KIRIM":
            del actuals[k]
    # Put entire income under "Oylik maosh" as default
    if income_total > 0:
        actuals[("KIRIM", "Oylik maosh")] = income_total

    # ── 5. Draw each category group ──────────────────────────
    for group_name, subcats in BUDGET_CATEGORIES.items():
        # Group header row
        ws.row_dimensions[current_row].height = 24
        _merge_and_style(
            ws, f"A{current_row}:I{current_row}",
            f"  {group_name}",
            FILL_HEADER,
            Font(name="Calibri", size=11, bold=True, color=WHITE),
            Alignment(horizontal="left", vertical="center", indent=1),
        )
        current_row += 1

        group_actual_sum = 0.0
        for idx, (sub_name, _code) in enumerate(subcats):
            row_fill = FILL_LIGHT if idx % 2 == 0 else FILL_CREAM
            ws.row_dimensions[current_row].height = 20

            # Name (A-C)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
            n = ws.cell(row=current_row, column=1)
            n.value = f"   {sub_name}"
            n.font = FONT_BODY
            n.alignment = LEFT
            n.fill = row_fill
            for c in range(2, 4):
                ws.cell(row=current_row, column=c).fill = row_fill

            # Plan (D-E) — user can fill later, we leave empty
            ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=5)
            p = ws.cell(row=current_row, column=4)
            p.value = None
            p.font = FONT_NUM
            p.alignment = RIGHT
            p.number_format = '#,##0'
            p.fill = row_fill
            ws.cell(row=current_row, column=5).fill = row_fill

            # Actual (F-G)
            actual = actuals.get((group_name, sub_name), 0)
            ws.merge_cells(start_row=current_row, start_column=6, end_row=current_row, end_column=7)
            a = ws.cell(row=current_row, column=6)
            a.value = actual if actual > 0 else None
            a.font = FONT_BOLD if actual > 0 else FONT_NUM
            a.alignment = RIGHT
            a.number_format = '#,##0'
            a.fill = row_fill
            ws.cell(row=current_row, column=7).fill = row_fill
            group_actual_sum += actual

            # Difference (H-I) — formula: Plan - Actual
            ws.merge_cells(start_row=current_row, start_column=8, end_row=current_row, end_column=9)
            d = ws.cell(row=current_row, column=8)
            d.value = f"=IFERROR(D{current_row}-F{current_row},0)"
            d.font = FONT_NUM
            d.alignment = RIGHT
            d.number_format = '#,##0;[Red]-#,##0'
            d.fill = row_fill
            ws.cell(row=current_row, column=9).fill = row_fill

            current_row += 1

        # Subtotal row for this group
        ws.row_dimensions[current_row].height = 22
        subtotal_fill = PatternFill("solid", fgColor=PEACH_LIGHT)
        for c in range(1, 10):
            ws.cell(row=current_row, column=c).fill = subtotal_fill

        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
        sub_label = ws.cell(row=current_row, column=1)
        sub_label.value = f"  Jami: {group_name}"
        sub_label.font = FONT_BOLD
        sub_label.alignment = LEFT

        ws.merge_cells(start_row=current_row, start_column=6, end_row=current_row, end_column=7)
        sub_val = ws.cell(row=current_row, column=6)
        sub_val.value = group_actual_sum
        sub_val.font = FONT_BOLD
        sub_val.alignment = RIGHT
        sub_val.number_format = '#,##0'

        current_row += 2  # Spacer

    # ── 6. RAW TRANSACTIONS LIST ─────────────────────────────
    current_row += 1
    ws.row_dimensions[current_row].height = 30
    _merge_and_style(
        ws, f"A{current_row}:I{current_row}",
        "  📋 BARCHA TRANZAKSIYALAR",
        FILL_HEADER, FONT_H2,
        Alignment(horizontal="left", vertical="center", indent=2),
    )
    current_row += 1

    headers = [("Sana", 2), ("Tur", 1), ("Kategoriya", 2), ("Summa", 2), ("Izoh", 2)]
    col_positions = []
    c = 1
    for label, span in headers:
        col_positions.append(c)
        if span > 1:
            ws.merge_cells(
                start_row=current_row, start_column=c,
                end_row=current_row, end_column=c + span - 1,
            )
        hcell = ws.cell(row=current_row, column=c)
        hcell.value = label
        hcell.fill = FILL_HEADER
        hcell.font = FONT_H2
        hcell.alignment = CENTER
        for fc in range(c, c + span):
            ws.cell(row=current_row, column=fc).fill = FILL_HEADER
        c += span
    current_row += 1

    if budget:
        for idx, log in enumerate(sorted(budget, key=lambda x: str(x.get("logged_date", "")), reverse=True)):
            row_fill = FILL_LIGHT if idx % 2 == 0 else FILL_CREAM
            ws.row_dimensions[current_row].height = 20

            # Sana (A-B)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
            ws.cell(row=current_row, column=1, value=str(log.get("logged_date", ""))).font = FONT_BODY
            ws.cell(row=current_row, column=1).alignment = LEFT
            for cc in range(1, 3):
                ws.cell(row=current_row, column=cc).fill = row_fill

            # Tur (C)
            is_income = log.get("type") == "income"
            tip_cell = ws.cell(row=current_row, column=3, value="💰 Kirim" if is_income else "💸 Chiqim")
            tip_cell.font = Font(name="Calibri", size=10, bold=True,
                                  color=GREEN if is_income else RED)
            tip_cell.alignment = CENTER
            tip_cell.fill = row_fill

            # Kategoriya (D-E)
            ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=5)
            cat_cell = ws.cell(row=current_row, column=4, value=log.get("category") or "—")
            cat_cell.font = FONT_BODY
            cat_cell.alignment = LEFT
            for cc in range(4, 6):
                ws.cell(row=current_row, column=cc).fill = row_fill

            # Summa (F-G)
            ws.merge_cells(start_row=current_row, start_column=6, end_row=current_row, end_column=7)
            amt_cell = ws.cell(row=current_row, column=6, value=float(log.get("amount") or 0))
            amt_cell.font = FONT_BOLD
            amt_cell.alignment = RIGHT
            amt_cell.number_format = '#,##0'
            for cc in range(6, 8):
                ws.cell(row=current_row, column=cc).fill = row_fill

            # Izoh (H-I)
            ws.merge_cells(start_row=current_row, start_column=8, end_row=current_row, end_column=9)
            note_cell = ws.cell(row=current_row, column=8, value=(log.get("note") or "")[:60])
            note_cell.font = FONT_SMALL
            note_cell.alignment = LEFT
            for cc in range(8, 10):
                ws.cell(row=current_row, column=cc).fill = row_fill

            current_row += 1
    else:
        ws.row_dimensions[current_row].height = 30
        _merge_and_style(
            ws, f"A{current_row}:I{current_row}",
            "📭 Ushbu davrda tranzaksiyalar yo'q.",
            FILL_LIGHT,
            Font(name="Calibri", size=10, italic=True, color=MUTED),
            CENTER,
        )
        current_row += 1

    # ── 7. COLUMN WIDTHS ──────────────────────────────────────
    widths = {"A": 8, "B": 8, "C": 12, "D": 10, "E": 10, "F": 10, "G": 10, "H": 15, "I": 15}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A5"


# ════════════════════════════════════════════════════════════════
# PUBLIC API
# ════════════════════════════════════════════════════════════════
async def generate_excel_report(
    user_id: int, start: date, end: date, currency: str = "UZS"
) -> str:
    """Generate the full Excel report and return the temp file path."""
    habits, budget = await asyncio.gather(
        db.get_habits_in_range(user_id, start, end),
        db.get_budget_in_range(user_id, start, end),
    )

    def _build() -> str:
        wb = Workbook()
        # Sheet 1 — Odatlar
        ws_habit = wb.active
        ws_habit.title = "Odatlar"
        _build_habit_sheet(ws_habit, habits, start, end)

        # Sheet 2 — Budjet
        ws_budget = wb.create_sheet("Budjet")
        _build_budget_sheet(ws_budget, budget, start, end, currency)

        tmp = tempfile.NamedTemporaryFile(
            delete=False,
            suffix=".xlsx",
            prefix=f"hisobot_{user_id}_{start.isoformat()}_{end.isoformat()}_",
        )
        wb.save(tmp.name)
        tmp.close()
        return tmp.name

    loop = asyncio.get_running_loop()
    path = await loop.run_in_executor(None, _build)
    logger.info("Excel report generated: user=%s path=%s", user_id, path)
    return path


def cleanup_file(path: str) -> None:
    try:
        if os.path.exists(path):
            os.remove(path)
    except OSError as e:
        logger.warning("Failed to delete file %s: %s", path, e)
